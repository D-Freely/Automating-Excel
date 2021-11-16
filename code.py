import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import (FuncFormatter)
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import formatting, styles, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import string

# USER INPUT SECTION

source_file_path = '(...)/test_data.xlsx'

# ----------------------------------------------------------------------------------------------------------------

pd.options.mode.chained_assignment = None

df = pd.read_excel(source_file_path, index_col=None)

df_fin = df.copy()
wrkstream_lst = df_fin['Workstream'].unique().tolist()
column_headers = df_fin.columns.tolist()

def removekey(d, key):
    r = dict(d)
    del r[key]
    return r

green_color = '99ffcc'
red_color = 'ffcccc'
black_color_font = '000000'
lighter_blue_color = '4472C4'
dark_blue_color = '44546A'
white_color = 'FFFFFF'
orange_color = 'ED7D31'
dark_green_color = '028844'

black_font = styles.Font(size=11, bold=False, color=black_color_font)
white_font = styles.Font(size=12, bold=False, color=white_color)
green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
lighter_blue_fill = styles.PatternFill(start_color=lighter_blue_color, end_color=lighter_blue_color, fill_type='solid')
dark_blue_fill = styles.PatternFill(start_color=dark_blue_color, end_color=dark_blue_color, fill_type='solid')
orange_fill = styles.PatternFill(start_color=orange_color, end_color=orange_color, fill_type='solid')
dark_green_fill = styles.PatternFill(start_color=dark_green_color, end_color=dark_blue_color, fill_type='solid')

# ---------------------------------------------------------------------------

for ws in wrkstream_lst:
    df_ws = df_fin[df_fin['Workstream'] == ws]
    file_name = ws + '.xlsx'
    pic_name = ws + '.png'

    variance_headers = []
    variance_colheads = []
    actual_headers = []
    actual_colheads = []
    forecast_colheads = []
    for header in column_headers:
        if 'variance' in header.lower():
            variance_headers.append(header)
            variance_colheads.append(df_ws[header])
        if 'actual' in header.lower():
            actual_headers.append(header)
            actual_colheads.append(df_ws[header])
        if 'forecast' in header.lower():
            forecast_colheads.append(df_ws[header])

    df_ws['Total Variance'] = sum(variance_colheads)

    x_points = [x.replace(' Actual', '') for x in actual_headers]

# ---------------------------------------------------------------------------

    actual_col_sum = [x.sum() for x in actual_colheads]
    forecast_col_sum = [x.sum() for x in forecast_colheads]
    x_axis = np.arange(len(x_points))
    fig, ax = plt.subplots()
    fig.set_size_inches(11, 7.5, forward=True)
    ax.ticklabel_format(useOffset=False, style='plain')
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: "£{:,}".format(x)))
    ax.bar(x_axis - 0.2, forecast_col_sum, 0.4, label='Forecast', color='#ff960d')
    ax.bar(x_axis + 0.2, actual_col_sum, 0.4, label = 'Actual', color='royalblue')
    ax.set_xticks(x_axis)
    ax.set_xticklabels(x_points)
    ax.legend()
    ax.set_ylabel('Amount (£)')
    ax.set_title('Forecast vs Actual')
    plt.savefig(pic_name)

# ---------------------------------------------------------------------------

    wb = Workbook()

    ws1 = wb.create_sheet('Actuals_Forecast_Rec')
    ws2 = wb.create_sheet('Bar Chart')
    wb.remove(wb['Sheet'])

    for r in dataframe_to_rows(df_ws, index=False, header=True):
        ws1.append(r)

    num_rows = len(df_ws.index) + 1

    ColNames = {}
    Current = 0
    for COL in ws1.iter_cols(1, ws1.max_column):
        ColNames[COL[0].value] = Current
        Current += 1

    var_col_range = {}
    var_col_let_num = {}
    for key, value in ColNames.items():
        if 'variance' in key.lower():
            var_col_range.update({get_column_letter(value+1) + str(2) : get_column_letter(value+1) + str(num_rows)})

    ColNames_minus_tvar = removekey(ColNames, 'Total Variance')

    for key, value in ColNames_minus_tvar.items():
        if 'variance' in key.lower():
            var_col_let_num.update({get_column_letter(value + 1) : value+1})

    for key, value in var_col_range.items():
        ws1.conditional_formatting.add(key + ':' + value, formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=green_fill, font=black_font))
        ws1.conditional_formatting.add(key + ':' + value, formatting.rule.CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill, font=black_font))

    ws1['A1'].fill = lighter_blue_fill
    ws1['B1'].fill = dark_blue_fill
    ws1['C1'].fill = dark_blue_fill
    alpha_lst = list(string.ascii_uppercase)
    format_range = alpha_lst[3:ColNames['Total Variance']+1]
    for letter in format_range[:-1]:
        ws1[letter + '1'].fill = dark_green_fill
    ws1[format_range[-1] + '1'].fill = orange_fill
    for cell in ws1["1:1"]:
        cell.font = white_font

    col_len_num_lst = list(range(num_rows+1))[1:]
    col_len_num_lst_2 = col_len_num_lst[1:]

    for letter in format_range:
        cell_lst = [letter + str(x) for x in col_len_num_lst]
        for cell in cell_lst:
            ws1[cell].number_format = '£#,##0.00'

    for let, num in var_col_let_num.items():
        cell_lst = [let + str(x) for x in col_len_num_lst_2]
        n = 0
        for cell in cell_lst:
            ws1[cell] = '=' + get_column_letter(num-1) + str(col_len_num_lst_2[n]) + '-' + get_column_letter(num-2) + str(col_len_num_lst_2[n])
            n += 1

    dim_holder = DimensionHolder(worksheet=ws1)
    for col in range(ws1.min_column, ws1.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws1, min=col, max=col, width=18)
    ws1.column_dimensions = dim_holder

    img = openpyxl.drawing.image.Image(pic_name)
    ws2.add_image(img)

    wb.save(file_name)

